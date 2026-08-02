from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


def dar_formato(ws):
    # Obtener las posiciones de las columnas según sus encabezados
    columnas = {
        celda.value: celda.column
        for celda in ws[1]
        if celda.value is not None
    }

    # Agregar columna Estado al final
    columna_estado = ws.max_column + 1
    ws.cell(
        row=1,
        column=columna_estado,
        value="Estado"
    )

    columnas["Estado"] = columna_estado

    # Formato de todos los encabezados
    for celda in ws[1]:
        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Congelar encabezados
    ws.freeze_panes = "A2"

    # Activar filtros
    ws.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(ws.max_column)}"
        f"{ws.max_row}"
    )

    # Ajustar automáticamente el ancho
    for columna in ws.columns:
        max_length = 0
        letra = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value is not None:
                max_length = max(
                    max_length,
                    len(str(celda.value))
                )

        ws.column_dimensions[letra].width = min(
            max_length + 4,
            70
        )

    # Anchos específicos
    anchos = {
        "Codigo": 20,
        "Nombre": 55,
        "Organismo": 45,
        "Presupuesto": 18,
        "Fecha de cierre": 21,
        "Ofertas": 10,
        "Fecha de actualización de ofertas": 28,
        "Ficha": 14,
        "Estado": 12
    }

    for nombre, ancho in anchos.items():
        if nombre in columnas:
            letra = get_column_letter(
                columnas[nombre]
            )

            ws.column_dimensions[letra].width = ancho

    # Si no hay filas de datos, terminamos
    if ws.max_row <= 1:
        return

    # Formato moneda para Presupuesto
    if "Presupuesto" in columnas:
        letra_presupuesto = get_column_letter(
            columnas["Presupuesto"]
        )

        for fila in range(2, ws.max_row + 1):
            ws[
                f"{letra_presupuesto}{fila}"
            ].number_format = '$#,##0'

    # Formato para las columnas de fecha
    columnas_fecha = [
        "Fecha de cierre",
        "Fecha de actualización de ofertas"
    ]

    for nombre in columnas_fecha:
        if nombre in columnas:
            letra_fecha = get_column_letter(
                columnas[nombre]
            )

            for fila in range(2, ws.max_row + 1):
                ws[
                    f"{letra_fecha}{fila}"
                ].number_format = "dd-mm-yyyy hh:mm"

    # Centrar columna Ofertas
    if "Ofertas" in columnas:
        letra_ofertas = get_column_letter(
            columnas["Ofertas"]
        )

        for fila in range(2, ws.max_row + 1):
            ws[
                f"{letra_ofertas}{fila}"
            ].alignment = Alignment(
                horizontal="center"
            )

    # Convertir la columna Ficha en hipervínculos
    if "Ficha" in columnas and "Codigo" in columnas:
        letra_ficha = get_column_letter(
            columnas["Ficha"]
        )

        letra_codigo = get_column_letter(
            columnas["Codigo"]
        )

        for fila in range(2, ws.max_row + 1):
            codigo = ws[
                f"{letra_codigo}{fila}"
            ].value

            celda_ficha = ws[
                f"{letra_ficha}{fila}"
            ]

            celda_ficha.hyperlink = (
                "https://buscador.mercadopublico.cl/"
                f"ficha?code={codigo}"
            )

            celda_ficha.value = "Revisar"
            celda_ficha.style = "Hyperlink"

    # Pintar de verde la fila cuando Estado tenga una x
    verde = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    letra_estado = get_column_letter(
        columnas["Estado"]
    )

    ultima_columna = get_column_letter(
        ws.max_column
    )

    regla = FormulaRule(
        formula=[
            f'${letra_estado}2="x"'
        ],
        fill=verde
    )

    ws.conditional_formatting.add(
        f"A2:{ultima_columna}{ws.max_row}",
        regla
    )